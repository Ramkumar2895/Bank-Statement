"""
Database operations.

TESTING MODE: Writes to XLSX files (same structure as MongoDB collections).
PRODUCTION: Set MONGODB_URI environment variable to use MongoDB for persistent cloud storage.

MongoDB collections structure:
  - statements: {_id, filename, uploaded_at, total_transactions, total_income, total_expense}
  - transactions: {_id, statement_id, date, description, category, debit, credit, balance, is_cash, saved_at}
  - passwords: {_id, bank_name, encrypted_password, created_at}
  - pending: {_pending_id, date, description, category, debit, credit, balance, is_cash, added_at}
"""
import os
import uuid
from datetime import datetime
from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Toggle: Auto-detect based on environment variable
# If MONGODB_URI is set, use MongoDB (cloud/production)
# Otherwise, fall back to XLSX (local/testing)
# ---------------------------------------------------------------------------
MONGODB_URI = os.getenv("MONGODB_URI", "").strip()
# USE_XLSX = not MONGODB_URI  # Use MongoDB if URI is provided, else use XLSX
USE_XLSX = False if MONGODB_URI else True
# Directory where XLSX files are stored (one file per collection)
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)

STATEMENTS_FILE = os.path.join(DATA_DIR, "statements.xlsx")
TRANSACTIONS_FILE = os.path.join(DATA_DIR, "transactions.xlsx")
PASSWORDS_FILE = os.path.join(DATA_DIR, "passwords.xlsx")
PENDING_FILE = os.path.join(DATA_DIR, "pending.xlsx")
KEY_FILE = os.path.join(DATA_DIR, ".encryption.key")

# Column definitions matching MongoDB document fields
STATEMENTS_COLUMNS = ["_id", "filename", "uploaded_at", "total_transactions", "total_income", "total_expense"]
TRANSACTIONS_COLUMNS = ["_id", "statement_id", "date", "description", "category", "debit", "credit", "balance", "is_cash", "saved_at"]
PASSWORDS_COLUMNS = ["_id", "bank_name", "encrypted_password", "created_at"]
PENDING_COLUMNS = ["_pending_id", "date", "description", "category", "debit", "credit", "balance", "is_cash", "added_at"]


# ---------------------------------------------------------------------------
# XLSX helper utilities
# ---------------------------------------------------------------------------

def _get_or_create_workbook(filepath: str, columns: list[str]) -> Workbook:
    """Load an existing workbook or create a new one with the given column headers."""
    if os.path.exists(filepath):
        return load_workbook(filepath)
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    wb.save(filepath)
    return wb


def _append_row(filepath: str, columns: list[str], doc: dict):
    """Append a single row to an XLSX file, creating it with headers if needed."""
    wb = _get_or_create_workbook(filepath, columns)
    ws = wb.active
    row = [doc.get(col, "") for col in columns]
    ws.append(row)
    wb.save(filepath)


def _append_rows(filepath: str, columns: list[str], docs: list[dict]):
    """Append multiple rows to an XLSX file, creating it with headers if needed."""
    if not docs:
        return
    wb = _get_or_create_workbook(filepath, columns)
    ws = wb.active
    for doc in docs:
        row = [doc.get(col, "") for col in columns]
        ws.append(row)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# MongoDB setup
# ---------------------------------------------------------------------------
_mongo_client = None
_mongo_db = None

def _get_mongo_db():
    """Connect to MongoDB and return the database object."""
    global _mongo_client, _mongo_db
    if _mongo_db is None:
        from pymongo import MongoClient
        _mongo_client = MongoClient(MONGODB_URI)
        _mongo_db = _mongo_client["bank_analyzer"]
        # Create indexes for faster queries
        _mongo_db.transactions.create_index([("statement_id", 1)])
        _mongo_db.transactions.create_index([("date", 1)])
        _mongo_db.passwords.create_index([("bank_name", 1)])
    return _mongo_db


def _get_existing_transaction_keys() -> set[tuple]:
    """Load existing (date, description, debit) keys from the transactions sheet."""
    keys = set()
    if USE_XLSX:
        if not os.path.exists(TRANSACTIONS_FILE):
            return keys
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        date_col = headers.index("date")
        desc_col = headers.index("description")
        debit_col = headers.index("debit")
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = str(row[date_col] or "").strip()
            desc_val = str(row[desc_col] or "").strip().lower()
            try:
                debit_val = round(float(row[debit_col] or 0), 2)
            except (ValueError, TypeError):
                debit_val = 0.0
            keys.add((date_val, desc_val, debit_val))
    else:
        db = _get_mongo_db()
        for doc in db.transactions.find({}, {"date": 1, "description": 1, "debit": 1}):
            keys.add((
                str(doc.get("date", "")).strip(),
                str(doc.get("description", "")).strip().lower(),
                round(float(doc.get("debit", 0)), 2),
            ))
    return keys


def get_saved_categories() -> dict[tuple, str]:
    """Load saved (date, description, debit) -> category map from the transactions sheet.

    When a user changes a category and saves, this lets us restore that choice
    on re-upload instead of using the auto-categorized value.
    """
    categories = {}
    if USE_XLSX:
        if not os.path.exists(TRANSACTIONS_FILE):
            return categories
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        date_col = headers.index("date")
        desc_col = headers.index("description")
        debit_col = headers.index("debit")
        cat_col = headers.index("category")
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = str(row[date_col] or "").strip()
            desc_val = str(row[desc_col] or "").strip().lower()
            try:
                debit_val = round(float(row[debit_col] or 0), 2)
            except (ValueError, TypeError):
                debit_val = 0.0
            cat_val = str(row[cat_col] or "").strip()
            if cat_val:
                categories[(date_val, desc_val, debit_val)] = cat_val
    else:
        db = _get_mongo_db()
        for doc in db.transactions.find({}, {"date": 1, "description": 1, "debit": 1, "category": 1}):
            date_val = str(doc.get("date", "")).strip()
            desc_val = str(doc.get("description", "")).strip().lower()
            debit_val = round(float(doc.get("debit", 0)), 2)
            cat_val = str(doc.get("category", "")).strip()
            if cat_val:
                categories[(date_val, desc_val, debit_val)] = cat_val
    return categories


def get_learned_categories() -> dict[str, str]:
    """Build a description-fragment -> category map from saved transactions.

    Extracts the meaningful merchant/person name from UPI/NEFT/ACH descriptions
    so that future transactions with the same payee auto-categorize correctly.
    e.g. 'UPI-HOTEL MANJUNATH BHAV-...' -> extract 'hotel manjunath' -> 'Food'
    """
    import re
    learned = {}
    if USE_XLSX:
        if not os.path.exists(TRANSACTIONS_FILE):
            return learned
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        desc_col = headers.index("description")
        cat_col = headers.index("category")
        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = str(row[desc_col] or "").strip()
            cat = str(row[cat_col] or "").strip()
            if not desc or not cat or cat == "Others":
                continue
            # Extract merchant/person name from common UPI/ACH patterns
            fragments = _extract_name_fragments(desc)
            for frag in fragments:
                if frag and len(frag) >= 3:
                    learned[frag] = cat
    else:
        db = _get_mongo_db()
        for doc in db.transactions.find({}, {"description": 1, "category": 1}):
            desc = str(doc.get("description", "")).strip()
            cat = str(doc.get("category", "")).strip()
            if not desc or not cat or cat == "Others":
                continue
            fragments = _extract_name_fragments(desc)
            for frag in fragments:
                if frag and len(frag) >= 3:
                    learned[frag] = cat
    return learned


def _extract_name_fragments(description: str) -> list[str]:
    """Extract searchable name fragments from a transaction description."""
    import re
    desc = description.lower().strip()
    fragments = []

    # UPI-MERCHANT NAME-... or UPI-PERSON NAME-...
    m = re.match(r'upi-(.+?)(?:-[a-z0-9@.]+-|-\d{10,})', desc)
    if m:
        name = m.group(1).strip()
        # Remove trailing reference junk
        name = re.sub(r'[a-z0-9]+@[a-z]+$', '', name).strip(' -')
        if name:
            fragments.append(name)
            # Also add first two words as a fragment for partial matching
            words = name.split()
            if len(words) >= 2:
                fragments.append(' '.join(words[:2]))

    # ACH D- MERCHANT-... 
    m2 = re.match(r'ach d-\s*(.+?)(?:-[a-z0-9]{8,})', desc)
    if m2:
        name = m2.group(1).strip()
        if name:
            fragments.append(name)

    # NEFT/IMPS patterns
    m3 = re.match(r'(?:neft|imps)[- ]+(.+?)(?:-[a-z0-9]{8,})', desc)
    if m3:
        name = m3.group(1).strip()
        if name:
            fragments.append(name)

    return fragments


# def get_all_transactions() -> list[dict]:
#     """Load all saved transactions from the database."""
#     transactions = []
#     if USE_XLSX:
#         if not os.path.exists(TRANSACTIONS_FILE):
#             return transactions
#         wb = load_workbook(TRANSACTIONS_FILE)
#         ws = wb.active
#         headers = [cell.value for cell in ws[1]]
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             txn = {}
#             for i, col in enumerate(headers):
#                 val = row[i]
#                 if col in ("debit", "credit", "balance"):
#                     try:
#                         val = round(float(val or 0), 2)
#                     except (ValueError, TypeError):
#                         val = 0.0
#                 elif col == "is_cash":
#                     val = str(val).upper() in ("TRUE", "1", "YES") if val else False
#                 else:
#                     val = str(val) if val else ""
#                 txn[col] = val
#             transactions.append(txn)
#     else:
#         db = _get_mongo_db()
#         for doc in db.transactions.find({}).sort("saved_at", -1):
#             txn = {
#                 "_id": str(doc.get("_id", "")),
#                 "statement_id": str(doc.get("statement_id", "")),
#                 "date": str(doc.get("date", "")),
#                 "description": str(doc.get("description", "")),
#                 "category": str(doc.get("category", "")),
#                 "debit": round(float(doc.get("debit", 0)), 2),
#                 "credit": round(float(doc.get("credit", 0)), 2),
#                 "balance": round(float(doc.get("balance", 0)), 2),
#                 "is_cash": bool(doc.get("is_cash", False)),
#                 "saved_at": str(doc.get("saved_at", "")),
#             }
#             transactions.append(txn)
#     return transactions

def get_all_transactions() -> list[dict]:
    """Load all saved transactions from the database (MongoDB or XLSX)."""
    transactions = []

    if USE_XLSX:
        if not os.path.exists(TRANSACTIONS_FILE):
            return transactions

        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            txn = {}

            for i, col in enumerate(headers):
                val = row[i]

                if col in ("debit", "credit", "balance"):
                    try:
                        val = round(float(val or 0), 2)
                    except (ValueError, TypeError):
                        val = 0.0

                elif col == "is_cash":
                    val = str(val).upper() in ("TRUE", "1", "YES") if val else False

                else:
                    val = str(val) if val else ""

                txn[col] = val

            # ✅ Ensure approved_date always exists
            if "approved_date" not in txn:
                txn["approved_date"] = txn.get("saved_at", "")

            transactions.append(txn)

    else:
        db = _get_mongo_db()

        for doc in db.transactions.find({}).sort("approved_date", -1):
            txn = {
                "_id": str(doc.get("_id", "")),
                "statement_id": str(doc.get("statement_id", "")),
                "date": str(doc.get("date", "")),
                "description": str(doc.get("description", "")),
                "category": str(doc.get("category", "")),

                "debit": round(float(doc.get("debit", 0)), 2),
                "credit": round(float(doc.get("credit", 0)), 2),
                "balance": round(float(doc.get("balance", 0)), 2),

                "is_cash": bool(doc.get("is_cash", False)),

                "saved_at": str(doc.get("saved_at", "")),

                # ✅ NEW FIELD (CRITICAL)
                "approved_date": (
                    doc.get("approved_date").isoformat()
                    if doc.get("approved_date")
                    else str(doc.get("saved_at", ""))
                )
            }

            transactions.append(txn)

    return transactions


def _get_existing_transaction_categories() -> dict[tuple, tuple]:
    """Load existing (date, description, debit) -> (row_number, category) from transactions sheet."""
    result = {}
    if USE_XLSX:
        if not os.path.exists(TRANSACTIONS_FILE):
            return result
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        date_col = headers.index("date")
        desc_col = headers.index("description")
        debit_col = headers.index("debit")
        cat_col = headers.index("category")
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            date_val = str(row[date_col] or "").strip()
            desc_val = str(row[desc_col] or "").strip().lower()
            try:
                debit_val = round(float(row[debit_col] or 0), 2)
            except (ValueError, TypeError):
                debit_val = 0.0
            cat_val = str(row[cat_col] or "").strip()
            result[(date_val, desc_val, debit_val)] = (row_num, cat_val)
    else:
        raise NotImplementedError("Set USE_XLSX = True or configure MongoDB")
    return result


def _update_transaction_categories(updates: list[tuple]):
    """Update category values in the transactions XLSX for given (row_number, new_category) pairs."""
    if not updates:
        return
    wb = load_workbook(TRANSACTIONS_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cat_col_idx = headers.index("category") + 1  # openpyxl is 1-indexed
    for row_num, new_category in updates:
        ws.cell(row=row_num, column=cat_col_idx, value=new_category)
    wb.save(TRANSACTIONS_FILE)


def update_transaction_category(date: str, description: str, debit: float, new_category: str) -> bool:
    """Update category for a single saved transaction identified by (date, description, debit). Returns True if updated."""
    if USE_XLSX:
        existing = _get_existing_transaction_categories()
        key = (date.strip(), description.strip().lower(), round(debit, 2))
        if key not in existing:
            return False
        row_num, old_cat = existing[key]
        if old_cat == new_category:
            return True  # already correct
        _update_transaction_categories([(row_num, new_category)])
        return True
    else:
        db = _get_mongo_db()
        result = db.transactions.update_one(
            {
                "date": date.strip(),
                "description": description.strip(),
                "debit": round(float(debit), 2),
            },
            {"$set": {"category": new_category}}
        )
        return result.modified_count > 0


def save_transactions(transactions: list[dict], filename: str) -> dict:
    """Save transactions after deduplication. Updates categories for existing ones. Returns info dict."""

    # Load existing transactions with their row numbers and categories
    existing = _get_existing_transaction_categories()
    new_transactions = []
    skipped = 0
    category_updates = []  # (row_number, new_category) for changed categories

    for t in transactions:
        key = (
            str(t["date"]).strip(),
            str(t["description"]).strip().lower(),
            round(float(t["debit"]), 2),
        )
        if key in existing:
            if USE_XLSX:
                row_num, saved_cat = existing[key]
                # If user changed the category, queue an update
                if t.get("category", "") and t["category"] != saved_cat:
                    category_updates.append((row_num, t["category"]))
            else:
                # For MongoDB, directly update in collection
                saved_cat = existing[key][1]
                if t.get("category", "") and t["category"] != saved_cat:
                    update_transaction_category(t["date"], t["description"], t["debit"], t["category"])
            skipped += 1
        else:
            new_transactions.append(t)
            existing[key] = (0, t.get("category", ""))  # avoid dupes within the same upload

    # Apply category updates to existing rows (XLSX only)
    updated = 0
    if category_updates and USE_XLSX:
        _update_transaction_categories(category_updates)
        updated = len(category_updates)
    elif not USE_XLSX and category_updates:
        updated = len(category_updates)

    if not new_transactions:
        return {
            "statement_id": None,
            "total_saved": 0,
            "skipped": skipped,
            "updated": updated,
        }

    statement_id = str(uuid.uuid4())

    statement_doc = {
        "_id": statement_id,
        "filename": filename,
        "uploaded_at": datetime.utcnow().isoformat(),
        "total_transactions": len(new_transactions),
        "total_income": round(sum(t["credit"] for t in new_transactions), 2),
        "total_expense": round(sum(t["debit"] for t in new_transactions), 2),
    }

    txn_docs = []
    for t in new_transactions:
        txn_docs.append({
            "_id": str(uuid.uuid4()),
            "statement_id": statement_id,
            "date": t["date"],
            "description": t["description"],
            "category": t["category"],
            "debit": t["debit"],
            "credit": t["credit"],
            "balance": t["balance"],
            "is_cash": t.get("is_cash", False),
            "saved_at": datetime.utcnow().isoformat(),
            "approved_date": datetime.now().isoformat(),
        })

    if USE_XLSX:
        _append_row(STATEMENTS_FILE, STATEMENTS_COLUMNS, statement_doc)
        if txn_docs:
            _append_rows(TRANSACTIONS_FILE, TRANSACTIONS_COLUMNS, txn_docs)
    else:
        db = _get_mongo_db()
        db.statements.insert_one(statement_doc)
        if txn_docs:
            db.transactions.insert_many(txn_docs)

    return {
        "statement_id": statement_id,
        "total_saved": len(new_transactions),
        "skipped": skipped,
        "updated": updated,
    }


# ---------------------------------------------------------------------------
# Password management (passwords collection)
# ---------------------------------------------------------------------------

# Default bank passwords — encrypted and stored in passwords.xlsx
_DEFAULT_PASSWORDS = {
    "HDFC": "168141974",
    "SBI": "RAMKU28111995",
}


def seed_passwords():
    """Seed default bank passwords into the passwords sheet (if not already present)."""
    existing = get_all_bank_names()
    for bank, pwd in _DEFAULT_PASSWORDS.items():
        if bank.upper() not in [b.upper() for b in existing]:
            save_bank_password(bank, pwd)


def save_bank_password(bank_name: str, password: str):
    """Encrypt and save a bank password."""
    doc = {
        "_id": str(uuid.uuid4()),
        "bank_name": bank_name.upper(),
        "encrypted_password": _encrypt(password),
        "created_at": datetime.utcnow().isoformat(),
    }
    if USE_XLSX:
        _append_row(PASSWORDS_FILE, PASSWORDS_COLUMNS, doc)
    else:
        # db = get_db()
        # db.passwords.insert_one(doc)
        raise NotImplementedError("Set USE_XLSX = True or configure MongoDB")


def get_bank_password(bank_name: str) -> str | None:
    """Retrieve and decrypt the password for a bank. Returns None if not found."""
    if USE_XLSX:
        if not os.path.exists(PASSWORDS_FILE):
            return None
        wb = load_workbook(PASSWORDS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        bank_col = headers.index("bank_name")
        pwd_col = headers.index("encrypted_password")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[bank_col] and row[bank_col].upper() == bank_name.upper():
                return _decrypt(row[pwd_col])
        return None
    else:
        # db = get_db()
        # doc = db.passwords.find_one({"bank_name": bank_name.upper()})
        # if doc:
        #     return _decrypt(doc["encrypted_password"])
        # return None
        raise NotImplementedError("Set USE_XLSX = True or configure MongoDB")


def get_all_bank_names() -> list[str]:
    """Return list of bank names that have stored passwords."""
    if USE_XLSX:
        if not os.path.exists(PASSWORDS_FILE):
            return []
        wb = load_workbook(PASSWORDS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        bank_col = headers.index("bank_name")
        return [row[bank_col] for row in ws.iter_rows(min_row=2, values_only=True) if row[bank_col]]
    else:
        raise NotImplementedError("Set USE_XLSX = True or configure MongoDB")


# ---------------------------------------------------------------------------
# Gmail configuration (encrypted, stored in passwords.xlsx as bank_name="GMAIL_CONFIG")
# ---------------------------------------------------------------------------

GMAIL_CONFIG_KEY = "GMAIL_CONFIG"


def save_gmail_config(gmail_user: str, gmail_app_password: str):
    """Save Gmail credentials encrypted in the passwords sheet."""
    if USE_XLSX:
        wb = _get_or_create_workbook(PASSWORDS_FILE, PASSWORDS_COLUMNS)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        bank_col = headers.index("bank_name")
        # Remove existing gmail config rows
        rows_to_delete = []
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=bank_col + 1).value == GMAIL_CONFIG_KEY:
                rows_to_delete.append(row_num)
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
        wb.save(PASSWORDS_FILE)
        # Save user and password as a single encrypted row
        _append_row(PASSWORDS_FILE, PASSWORDS_COLUMNS, {
            "_id": str(uuid.uuid4()),
            "bank_name": GMAIL_CONFIG_KEY,
            "encrypted_password": _encrypt(f"{gmail_user}|||{gmail_app_password}"),
            "created_at": datetime.utcnow().isoformat(),
        })


def get_gmail_config() -> dict | None:
    """Retrieve saved Gmail credentials. Returns {gmail_user, gmail_password} or None."""
    if USE_XLSX:
        if not os.path.exists(PASSWORDS_FILE):
            return None
        wb = load_workbook(PASSWORDS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        bank_col = headers.index("bank_name")
        pwd_col = headers.index("encrypted_password")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[bank_col] == GMAIL_CONFIG_KEY:
                decrypted = _decrypt(row[pwd_col])
                parts = decrypted.split("|||", 1)
                if len(parts) == 2:
                    return {"gmail_user": parts[0], "gmail_password": parts[1]}
        return None


def delete_gmail_config():
    """Remove saved Gmail credentials."""
    if USE_XLSX:
        if not os.path.exists(PASSWORDS_FILE):
            return
        wb = load_workbook(PASSWORDS_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        bank_col = headers.index("bank_name")
        rows_to_delete = []
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=bank_col + 1).value == GMAIL_CONFIG_KEY:
                rows_to_delete.append(row_num)
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
        wb.save(PASSWORDS_FILE)


# ---------------------------------------------------------------------------
# Pending transactions persistence
# ---------------------------------------------------------------------------

def load_pending_transactions() -> list[dict]:
    """Load all pending transactions from the database."""
    pending = []
    if USE_XLSX:
        if not os.path.exists(PENDING_FILE):
            return pending
        wb = load_workbook(PENDING_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            txn = {}
            for i, col in enumerate(headers):
                val = row[i]
                if col in ("debit", "credit", "balance"):
                    try:
                        val = round(float(val or 0), 2)
                    except (ValueError, TypeError):
                        val = 0.0
                elif col == "is_cash":
                    val = str(val).upper() in ("TRUE", "1", "YES") if val else False
                else:
                    val = str(val) if val else ""
                txn[col] = val
            pending.append(txn)
    else:
        db = _get_mongo_db()
        for doc in db.pending.find({}).sort("added_at", -1):
            txn = {
                "_pending_id": str(doc.get("_pending_id", "")),
                "date": str(doc.get("date", "")),
                "description": str(doc.get("description", "")),
                "category": str(doc.get("category", "")),
                "debit": round(float(doc.get("debit", 0)), 2),
                "credit": round(float(doc.get("credit", 0)), 2),
                "balance": round(float(doc.get("balance", 0)), 2),
                "is_cash": bool(doc.get("is_cash", False)),
                "added_at": str(doc.get("added_at", "")),
            }
            pending.append(txn)
    return pending


def save_pending_transactions(pending: list[dict]):
    """Save pending transactions to the database."""
    if USE_XLSX:
        wb = Workbook()
        ws = wb.active
        ws.append(PENDING_COLUMNS)
        for t in pending:
            row = [t.get(col, "") for col in PENDING_COLUMNS]
            ws.append(row)
        wb.save(PENDING_FILE)
    else:
        db = _get_mongo_db()
        # Clear existing pending transactions
        db.pending.delete_many({})
        # Insert all pending transactions
        if pending:
            docs_to_insert = []
            for t in pending:
                docs_to_insert.append({
                    "_pending_id": t.get("_pending_id", str(uuid.uuid4())),
                    "date": t.get("date", ""),
                    "description": t.get("description", ""),
                    "category": t.get("category", ""),
                    "debit": float(t.get("debit", 0)),
                    "credit": float(t.get("credit", 0)),
                    "balance": float(t.get("balance", 0)),
                    "is_cash": bool(t.get("is_cash", False)),
                    "added_at": t.get("added_at", datetime.utcnow().isoformat()),
                })
            db.pending.insert_many(docs_to_insert)


def clear_pending_transactions():
    """Remove all pending transactions."""
    if USE_XLSX:
        if os.path.exists(PENDING_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(PENDING_COLUMNS)
            wb.save(PENDING_FILE)
    else:
        db = _get_mongo_db()
        db.pending.delete_many({})